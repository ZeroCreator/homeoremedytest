"""
Модуль для экспорта данных в Excel с использованием openpyxl
"""
import json
from datetime import datetime
from pathlib import Path
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Класс для экспорта данных в Excel с использованием openpyxl"""

    # Константы для форматирования
    HEADER_COLOR = '4A6FA5'      # Синий цвет заголовков
    FONT_NAME = 'Calibri'
    FONT_SIZE = 11
    HEADER_FONT_SIZE = 12

    # Цвета для сложности
    DIFFICULTY_COLORS = {
        'Легкий': 'E8F5E8',      # Светло-зеленый
        'Средний': 'FFF3E0',     # Светло-оранжевый
        'Сложный': 'FFEBEE'      # Светло-красный
    }

    # Цвет для скрытых карточек (серый фон)
    HIDDEN_COLOR = 'F5F5F5'      # Светло-серый

    # Коэффициенты для расчета ширины колонок
    CHAR_WIDTH = 1.2  # Ширина одного символа
    MAX_COLUMN_WIDTH = 50
    MIN_COLUMN_WIDTH = 5

    def __init__(self, json_file_path=None):
        """
        Инициализация экспортера

        Args:
            json_file_path: Путь к JSON файлу с данными
        """
        if json_file_path:
            self.json_file_path = Path(json_file_path)
        else:
            from config import Config
            self.json_file_path = Config.JSON_FILE

    def load_cards(self):
        """Загрузка карточек из JSON файла"""
        try:
            if self.json_file_path.exists():
                with open(self.json_file_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            print(f"Ошибка загрузки JSON: {e}")

        return {"cards": [], "themes": [], "next_id": 1}

    def export_to_excel(self):
        """
        Создание Excel файла с карточками

        Returns:
            BytesIO: Буфер с Excel файлом
            str: Имя файла для скачивания
        """
        try:
            # Загружаем данные
            cards_data = self.load_cards()
            cards = cards_data.get('cards', [])

            if not cards:
                raise ValueError("Нет данных для экспорта")

            print(f"Экспорт {len(cards)} карточек...")

            # Создаем рабочую книгу Excel
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Карточки'

            # Заголовки колонок (добавлена колонка "Версия")
            headers = ['№', 'Вопрос', 'Ответ', 'Объяснение', 'Тема', 'Сложность', 'Скрытый', 'Версия']
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)

                # Стиль заголовка
                cell.font = Font(name=self.FONT_NAME, size=self.HEADER_FONT_SIZE,
                               bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.HEADER_COLOR,
                                      end_color=self.HEADER_COLOR,
                                      fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Границы
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border

            # Маппинг сложности
            difficulty_map = {
                'easy': 'Легкий',
                'medium': 'Средний',
                'hard': 'Сложный'
            }

            # Заполняем данными
            for row_idx, card in enumerate(cards, start=2):
                # Преобразуем данные
                difficulty_text = difficulty_map.get(card.get('difficulty', 'medium'), 'Средний')
                hidden_text = 'Да' if card.get('hidden', False) else 'Нет'
                version_text = card.get('version', '')  # Получаем версию, может быть пустой

                # Очищаем текст от лишних символов
                def clean_cell_text(text):
                    if not text:
                        return ""
                    # Убираем Windows символы
                    text = str(text).replace('_x000D_', '\n')
                    text = text.replace('\r\n', '\n')
                    text = text.replace('\r', '\n')
                    return text.strip()

                # Данные для записи (добавлена версия)
                data = [
                    card['id'],
                    clean_cell_text(card['question']),
                    clean_cell_text(card['answer']),
                    clean_cell_text(card.get('explanation', '')),
                    clean_cell_text(card['theme']),
                    difficulty_text,
                    hidden_text,
                    clean_cell_text(version_text)
                ]

                # Определяем цвет фона в зависимости от сложности
                bg_color = self.DIFFICULTY_COLORS.get(difficulty_text, 'FFFFFF')

                # Если карточка скрыта, используем серый фон
                if card.get('hidden', False):
                    bg_color = self.HIDDEN_COLOR

                for col_idx, value in enumerate(data, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

                    # Базовый стиль
                    cell.font = Font(name=self.FONT_NAME, size=self.FONT_SIZE)
                    cell.border = thin_border
                    cell.alignment = Alignment(
                        vertical='top',
                        wrap_text=True
                    )

                    # Центрирование для определенных колонок
                    if col_idx in [1, 6, 7, 8]:  # №, Сложность, Скрытый, Версия
                        cell.alignment = Alignment(
                            horizontal='center',
                            vertical='center',
                            wrap_text=True
                        )
                    else:
                        cell.alignment = Alignment(
                            horizontal='left',
                            vertical='top',
                            wrap_text=True
                        )

                    # Фон для ячейки
                    if bg_color != 'FFFFFF':
                        cell.fill = PatternFill(start_color=bg_color,
                                              end_color=bg_color,
                                              fill_type='solid')

            # Автоматическая настройка ширины колонок
            self._adjust_column_widths(worksheet)

            # Настраиваем высоту строк
            self._adjust_row_heights(worksheet)

            # Настраиваем фильтры (автофильтр)
            worksheet.auto_filter.ref = worksheet.dimensions

            # Замораживаем заголовки
            worksheet.freeze_panes = 'A2'

            # Сохраняем в буфер
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            # Генерируем имя файла
            filename = self._generate_filename()

            print(f"Экспорт завершен: {len(cards)} карточек")
            return buffer, filename

        except Exception as e:
            print(f"Ошибка при экспорте в Excel: {e}")
            import traceback
            traceback.print_exc()
            raise

    def _adjust_column_widths(self, worksheet):
        """Автоматическая настройка ширины колонок"""
        column_widths = []

        # Устанавливаем фиксированные ширины для последних колонок
        fixed_widths = {
            'F': 15,  # Сложность
            'G': 15,  # Скрытый
            'H': 12  # Версия
        }

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            # Если для этой колонки есть фиксированная ширина
            if column_letter in fixed_widths:
                worksheet.column_dimensions[column_letter].width = fixed_widths[column_letter]
                column_widths.append(fixed_widths[column_letter])
                continue

            # Для остальных колонок вычисляем оптимальную ширину
            for cell in column:
                try:
                    if cell.value:
                        # Учитываем заголовки (первая строка) отдельно
                        if cell.row == 1:  # Заголовок
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                        else:  # Данные
                            lines = str(cell.value).split('\n')
                            max_line_length = max(len(line) for line in lines) if lines else 0
                            max_length = max(max_length, max_line_length)
                except:
                    pass

            # Рассчитываем ширину с учетом минимальных значений
            min_widths = {
                'A': 6,  # №
                'B': 15,  # Вопрос
                'C': 15,  # Ответ
                'D': 15,  # Объяснение
                'E': 12  # Тема
            }

            min_width = min_widths.get(column_letter, 8)
            adjusted_width = min(
                self.MAX_COLUMN_WIDTH,
                max(min_width, (max_length + 2) * self.CHAR_WIDTH)
            )

            worksheet.column_dimensions[column_letter].width = adjusted_width
            column_widths.append(adjusted_width)

            # Для отладки
            header_text = worksheet.cell(row=1, column=column[0].column).value
            print(f"Колонка {column_letter} ('{header_text}'): ширина {adjusted_width:.1f}")

        return column_widths

    def _adjust_row_heights(self, worksheet):
        """Настройка высоты строк для лучшего отображения текста"""
        for row in worksheet.iter_rows(min_row=2):  # Пропускаем заголовки
            max_lines = 1

            # Считаем максимальное количество строк в ячейках этой строки
            for cell in row:
                if cell.value and cell.column in [2, 3, 4, 5, 6]:  # Текст в колонках B, C, D, E, F (вопрос, ответ, объяснение, тема, версия)
                    # Считаем строки с учетом переносов
                    lines = str(cell.value).count('\n') + 1
                    max_lines = max(max_lines, lines)

            # Устанавливаем высоту строки (примерно 15 пикселей на строку)
            if max_lines > 1:
                height = min(100, 15 * max_lines)  # Максимум 100 пикселей
                worksheet.row_dimensions[row[0].row].height = height

        # Устанавливаем фиксированную высоту для заголовка
        worksheet.row_dimensions[1].height = 30

    def _generate_filename(self):
        """Генерация имени файла с датой"""
        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        return f"homeopathy_cards_{date_str}.xlsx"

    def export_to_file(self, output_path=None):
        """
        Экспорт в файл на диск

        Args:
            output_path: Путь для сохранения файла

        Returns:
            str: Путь к сохраненному файлу
        """
        if not output_path:
            output_path = Path.cwd() / self._generate_filename()

        buffer, _ = self.export_to_excel()

        with open(output_path, 'wb') as f:
            f.write(buffer.getvalue())

        print(f"Файл сохранен: {output_path}")
        return str(output_path)


# Фабричная функция
def create_exporter(json_file_path=None):
    """Создание экземпляра экспортера"""
    return ExcelExporter(json_file_path)


# Тестовые функции
def test_exporter():
    """Тестирование экспортера"""
    print("Тестирование экспорта в Excel с openpyxl...")

    exporter = create_exporter("app/data/test_cards.json")

    try:
        # Экспорт в файл
        file_path = exporter.export_to_file("test_openpyxl.xlsx")
        print(f"✅ Файл успешно создан: {file_path}")

        # Проверяем размер файла
        import os
        file_size = os.path.getsize(file_path) / 1024
        print(f"📊 Размер файла: {file_size:.2f} KB")

        # Читаем файл для проверки
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        print(f"📝 Лист: {ws.title}")
        print(f"📊 Строк: {ws.max_row}")
        print(f"📊 Колонок: {ws.max_column}")

        # Проверяем первые 3 строки
        print("\nПервые 3 строки:")
        for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
            print(row)

        wb.close()
        return True

    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return False


def create_test_data():
    """Создание тестовых данных с версиями"""
    test_data = {
        "cards": [
            {
                "id": 1,
                "question": "Что такое гомеопатия?",
                "answer": "Метод лечения, основанный на принципе 'подобное лечится подобным'.",
                "explanation": "Создана Самуэлем Ганеманом в конце 18 века.",
                "theme": "Основы",
                "version": "Тест 1",
                "difficulty": "easy",
                "hidden": False
            },
            {
                "id": 2,
                "question": "Основные принципы гомеопатии_x000D_Второй принцип",
                "answer": "1. Принцип подобия_x000D_2. Принцип минимальной дозы",
                "explanation": "Эти принципы отличают гомеопатию от аллопатии.",
                "theme": "Принципы",
                "version": "Тест 2",
                "difficulty": "medium",
                "hidden": True  # Скрытая карточка
            }
        ],
        "themes": ["Основы", "Принципы"],
        "next_id": 3
    }

    # Сохраняем тестовые данные
    test_json = "app/data/test_cards.json"
    Path("app/data").mkdir(parents=True, exist_ok=True)

    with open(test_json, 'w', encoding='utf-8') as f:
        json.dump(test_data, f, ensure_ascii=False, indent=2)

    print(f"Созданы тестовые данные в {test_json}")
    return test_json


if __name__ == "__main__":
    print("=" * 60)
    print("ТЕСТИРОВАНИЕ ЭКСПОРТЕРА НА OPENPYXL")
    print("=" * 60)

    # Создаем тестовые данные
    test_json = create_test_data()

    # Запускаем тест
    if test_exporter():
        print("\n" + "=" * 60)
        print("✅ Все тесты пройдены успешно!")
        print("=" * 60)
    else:
        print("\n" + "=" * 60)
        print("❌ Тестирование завершилось с ошибками")
        print("=" * 60)
