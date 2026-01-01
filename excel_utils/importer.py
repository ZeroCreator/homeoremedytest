"""
Модуль для импорта данных из Excel файлов
Использует openpyxl для чтения файлов
Фиксированный формат колонок:
0: №,
1: Вопрос,
2: Ответ,
3: Объяснение,
4: Тема,
5: Сложность,
6: Скрытый
7: Версия,
"""
import json
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook


class ExcelImporter:
    """Класс для импорта данных из Excel файлов с фиксированным форматом"""

    def __init__(self, json_file_path=None):
        """
        Инициализация импортера
        Args:
            json_file_path: Путь к JSON файлу с данными
        """
        if json_file_path:
            self.json_file_path = Path(json_file_path)
        else:
            from config import Config
            self.json_file_path = Config.JSON_FILE

    def clean_text(self, text):
        """Очистка текста от Windows-символов и лишних пробелов"""
        if not text:
            return ""

        # Преобразуем в строку
        text = str(text)

        # Убираем Windows символы переноса строки
        text = text.replace('_x000D_', '\n')
        text = text.replace('\r\n', '\n')  # Windows перенос
        text = text.replace('\r', '\n')    # Mac старый перенос

        # Убираем лишние пробелы в начале и конце каждой строки
        lines = [line.strip() for line in text.split('\n')]

        # Убираем полностью пустые строки
        lines = [line for line in lines if line]

        return '\n'.join(lines)

    def normalize_text(self, text):
        """Нормализация текста для сравнения (поиск дубликатов)"""
        cleaned = self.clean_text(text)
        if not cleaned:
            return ""
        # Приводим к нижнему регистру, убираем пунктуацию
        text_lower = cleaned.lower()
        text_lower = re.sub(r'[^\w\s]', '', text_lower)
        text_lower = re.sub(r'\s+', ' ', text_lower)
        return text_lower.strip()

    def load_current_data(self):
        """Загрузка текущих данных из JSON"""
        try:
            if self.json_file_path.exists():
                with open(self.json_file_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            print(f"Ошибка загрузки JSON: {e}")

        return {"cards": [], "themes": [], "next_id": 1}

    def save_data(self, data):
        """Сохранение данных в JSON"""
        try:
            self.json_file_path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.json_file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"Ошибка сохранения JSON: {e}")
            return False

    def validate_excel_file(self, file_path):
        """
        Валидация Excel файла
        Args:
            file_path: Путь к Excel файлу
        Returns:
            tuple: (bool, str) - успех и сообщение
        """
        try:
            file_path = Path(file_path)

            # Проверка существования файла
            if not file_path.exists():
                return False, "Файл не существует"

            # Проверка расширения
            if file_path.suffix.lower() not in ['.xlsx', '.xls']:
                return False, "Поддерживаются только файлы Excel (.xlsx, .xls)"

            # Проверка размера
            file_size = file_path.stat().st_size
            if file_size == 0:
                return False, "Файл пустой"
            if file_size > 50 * 1024 * 1024:  # 50MB
                return False, "Файл слишком большой (максимум 50MB)"

            # Попытка открыть файл
            try:
                wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
                wb.close()
            except Exception as e:
                return False, f"Не удалось открыть файл: {str(e)}"

            return True, "Файл прошел проверку"

        except Exception as e:
            return False, f"Ошибка при проверке файла: {str(e)}"

    def read_excel_file(self, file_path, max_rows=10000):
        """
        Чтение данных из Excel файла
        Args:
            file_path: Путь к Excel файлу
            max_rows: Максимальное количество строк для чтения (включая заголовки)
        Returns:
            list: Список строк с данными
        """
        try:
            wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)

            # Определяем лист для чтения
            sheet_names = wb.sheetnames
            if not sheet_names:
                wb.close()
                return []

            # Используем первый лист
            ws = wb[sheet_names[0]]

            # Читаем данные
            data = []
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i > max_rows:
                    break

                # Преобразуем кортеж в список
                row_data = list(row)

                # Обрабатываем каждую ячейку
                for j, cell in enumerate(row_data):
                    if cell is None:
                        row_data[j] = ''
                    elif isinstance(cell, datetime):
                        row_data[j] = cell.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        # Очищаем текст сразу при чтении
                        row_data[j] = self.clean_text(cell)

                data.append(row_data)

            wb.close()
            return data

        except Exception as e:
            raise Exception(f"Ошибка чтения Excel файла: {str(e)}")

    def parse_excel_data(self, excel_data):
        """
        Парсинг данных из Excel с фиксированными колонками
        Args:
            excel_data: Данные из Excel файла
        Returns:
            list: Список карточек
        """
        if not excel_data or len(excel_data) < 2:
            return []

        # Проверяем заголовки (опционально, для информативности)
        headers = excel_data[0]
        expected_headers = ['№', 'Вопрос', 'Ответ', 'Объяснение', 'Тема', 'Сложность', 'Скрытый', 'Версия']

        print(f"Заголовки в файле: {headers[:8]}")
        print(f"Ожидаемые заголовки: {expected_headers}")

        # Проверяем соответствие заголовков
        for i in range(min(len(headers), len(expected_headers))):
            actual = str(headers[i]) if i < len(headers) else "(отсутствует)"
            expected = expected_headers[i]
            if actual != expected:
                print(f"⚠️  Внимание: колонка {i+1} - ожидалось '{expected}', найдено '{actual}'")

        # Маппинг значений сложности
        difficulty_map = {
            'легкий': 'easy', 'easy': 'easy', 'Легкий': 'easy',
            'средний': 'medium', 'medium': 'medium', 'Средний': 'medium',
            'сложный': 'hard', 'hard': 'hard', 'Сложный': 'hard'
        }

        # Маппинг значений скрытия
        hidden_map = {
            'да': True, 'yes': True, 'true': True, '1': True, '+': True, 'Да': True,
            'нет': False, 'no': False, 'false': False, '0': False, '-': False, 'Нет': False
        }

        # Парсим данные
        cards = []
        for row_idx, row in enumerate(excel_data[1:], start=2):
            try:
                # Фиксированные колонки (согласно экспортеру):
                # 0: №, 1: Вопрос, 2: Ответ, 3: Объяснение, 4: Тема, 5: Сложность, 6: Скрытый, 7: Версия

                # Проверяем, что строка имеет достаточно колонок
                if len(row) < 8:
                    # Дополняем пустыми значениями
                    row = list(row) + [''] * (8 - len(row))

                # Извлекаем значения (уже очищенные при чтении)
                id_str = row[0] if len(row) > 0 else ''
                question = row[1] if len(row) > 1 else ''
                answer = row[2] if len(row) > 2 else ''
                explanation = row[3] if len(row) > 3 else ''
                theme = row[4] if len(row) > 4 else ''
                difficulty_str = row[5] if len(row) > 5 else ''
                hidden_str = row[6] if len(row) > 6 else ''
                version = row[7] if len(row) > 7 else ''

                # Проверяем обязательные поля
                if not question and not answer:
                    print(f"DEBUG: Строка {row_idx} пропущена - пустые вопрос и ответ")
                    continue

                # ID (опционально)
                try:
                    card_id = int(id_str) if id_str and id_str.isdigit() else 0
                except ValueError:
                    card_id = 0

                # Сложность
                difficulty_normalized = difficulty_str.lower()
                difficulty = difficulty_map.get(difficulty_normalized, 'medium')

                # Скрытый
                hidden_normalized = hidden_str.lower()
                hidden = hidden_map.get(hidden_normalized, False)

                # Создаем карточку с версией
                card = {
                    'id': card_id,
                    'question': question,
                    'answer': answer,
                    'explanation': explanation,
                    'theme': theme,
                    'difficulty': difficulty,
                    'hidden': hidden,
                    'version': version
                }

                cards.append(card)

                # Отладочный вывод
                if row_idx <= 5 or row_idx >= len(excel_data) - 5:
                    question_preview = question[:50] + '...' if len(question) > 50 else question
                    version_preview = f", Версия='{version}'" if version else ""
                    print(f"Строка {row_idx}: ID={card_id}, Вопрос='{question_preview}'{version_preview}")

            except Exception as e:
                print(f"Ошибка в строке {row_idx}: {str(e)}")
                continue

        return cards

    def import_from_excel(self, excel_file_path, mode='append'):
        """
        Основная функция импорта
        Args:
            excel_file_path: Путь к Excel файлу
            mode: Режим импорта ('append' или 'replace')
        Returns:
            tuple: (bool, dict) - успех и статистика
        """
        try:
            print(f"\n{'='*60}")
            print(f"НАЧАЛО ИМПОРТА: файл={excel_file_path}, режим={mode}")
            print(f"{'='*60}")

            # Загружаем текущие данные
            current_data = self.load_current_data()
            current_cards = current_data.get('cards', [])
            current_themes = set(current_data.get('themes', []))
            next_id = current_data.get('next_id', 1)

            print(f"Текущая БД: {len(current_cards)} карточек, next_id={next_id}")

            # Читаем Excel файл
            excel_data = self.read_excel_file(excel_file_path, max_rows=1000)

            if not excel_data:
                return False, {'error': 'Файл не содержит данных или пуст'}

            print(f"Прочитано строк из файла: {len(excel_data)} (включая заголовки)")

            # Парсим данные (фиксированный формат)
            imported_cards = self.parse_excel_data(excel_data)

            if not imported_cards:
                return False, {'error': 'Не удалось извлечь данные из файла'}

            print(f"Найдено карточек в файле: {len(imported_cards)}")

            # Обрабатываем режим импорта
            if mode == 'replace':
                all_cards = imported_cards
                current_themes = set()
                imported_count = len(imported_cards)
                skipped_count = 0
                errors_count = 0

                print(f"Режим 'replace': все {imported_count} карточек будут заменены")

            else:  # append
                # Создаем словарь существующих вопросов для проверки дубликатов
                existing_questions = {}
                for card in current_cards:
                    question_key = self.normalize_text(card['question'])
                    existing_questions[question_key] = {
                        'id': card['id'],
                        'original': card['question']
                    }

                print(f"Создан словарь существующих вопросов: {len(existing_questions)} ключей")

                all_cards = current_cards.copy()
                imported_count = 0
                skipped_count = 0
                errors_count = 0

                # Проверяем каждую импортируемую карточку
                for card_idx, card in enumerate(imported_cards, 1):
                    try:
                        question_key = self.normalize_text(card['question'])
                        question_display = card['question'][:50] + '...' if len(card['question']) > 50 else card['question']

                        if question_key in existing_questions:
                            existing = existing_questions[question_key]
                            skipped_count += 1
                            if skipped_count <= 3 or card_idx >= len(imported_cards) - 2:
                                print(f"Дубликат #{card_idx}: '{question_display}'")
                                print(f"  Уже есть ID {existing['id']}: '{existing['original'][:50]}...'")
                        else:
                            imported_count += 1
                            existing_questions[question_key] = {
                                'id': card['id'],
                                'original': card['question']
                            }
                            all_cards.append(card)
                            if imported_count <= 3 or card_idx >= len(imported_cards) - 2:
                                print(f"Новая #{card_idx}: '{question_display}'")
                    except Exception as e:
                        errors_count += 1
                        print(f"Ошибка при обработке карточки #{card_idx}: {str(e)}")

            # Извлекаем темы из всех карточек
            all_themes = set(current_themes)  # Начинаем с текущих тем
            for card in all_cards:
                if card['theme']:
                    # Тема может содержать несколько значений через запятую
                    themes = [t.strip() for t in str(card['theme']).split(',')]
                    for theme in themes:
                        if theme:
                            all_themes.add(theme)

            # Пересчитываем ID если нужно
            if mode == 'replace' or not current_cards:
                print(f"Пересчет ID для всех карточек...")
                for i, card in enumerate(all_cards, start=1):
                    card['id'] = i
                next_id = len(all_cards) + 1
            else:
                # Находим максимальный ID среди всех карточек
                max_id = 0
                for card in all_cards:
                    if card['id'] > max_id:
                        max_id = card['id']

                # Присваиваем ID новым карточкам (если у них ID=0 или конфликтуют)
                new_id = max_id + 1
                assigned_ids = set(card['id'] for card in all_cards[:len(current_cards)])

                for card in all_cards[len(current_cards):]:  # Только новые карточки
                    if card['id'] == 0 or card['id'] in assigned_ids:
                        # Назначаем новый ID
                        while new_id in assigned_ids:
                            new_id += 1
                        card['id'] = new_id
                        assigned_ids.add(new_id)
                        new_id += 1
                    else:
                        assigned_ids.add(card['id'])

                next_id = new_id

            # Сортируем по ID
            all_cards.sort(key=lambda x: x['id'])

            # Сохраняем данные
            updated_data = {
                'cards': all_cards,
                'themes': sorted(list(all_themes)),
                'next_id': next_id
            }

            print(f"\nСохранение данных:")
            print(f"  Карточек: {len(all_cards)} (было {len(current_cards)})")
            print(f"  Тем: {len(all_themes)} (было {len(current_themes)})")
            print(f"  Следующий ID: {next_id} (было {current_data.get('next_id', 1)})")

            if not self.save_data(updated_data):
                return False, {'error': 'Ошибка сохранения данных'}

            # Формируем статистику
            stats = {
                'imported': imported_count,
                'skipped': skipped_count,
                'errors': errors_count,
                'total': len(all_cards),
                'themes': len(all_themes),
                'mode': mode,
                'next_id': next_id,
                'previous_total': len(current_cards)
            }

            print(f"\n{'='*60}")
            print(f"ИТОГО ИМПОРТА:")
            print(f"  Добавлено новых: {imported_count}")
            print(f"  Пропущено (дубликаты): {skipped_count}")
            print(f"  Ошибок: {errors_count}")
            print(f"  Всего в БД: {len(all_cards)} карточек (было {len(current_cards)})")
            print(f"  Изменение: {len(all_cards) - len(current_cards):+d}")
            print(f"  Тем: {len(all_themes)}")
            print(f"  Следующий ID: {next_id}")
            print(f"{'='*60}")

            return True, stats

        except Exception as e:
            print(f"Ошибка импорта: {str(e)}")
            import traceback
            traceback.print_exc()
            return False, {'error': f'Ошибка импорта: {str(e)}'}

    def get_import_preview(self, file_path, limit=1000):
        """
        Предпросмотр данных из файла
        Args:
            file_path: Путь к файлу
            limit: Максимальное количество строк для показа
        """
        try:
            # Валидируем файл
            is_valid, message = self.validate_excel_file(file_path)
            if not is_valid:
                return False, message

            # Читаем файл
            excel_data = self.read_excel_file(file_path, max_rows=limit + 10)

            if not excel_data:
                return False, "Файл не содержит данных"

            # Фиксированные заголовки (согласно экспортеру)
            fixed_headers = ['№', 'Вопрос', 'Ответ', 'Объяснение', 'Тема', 'Сложность', 'Скрытый', 'Версия']

            # Проверяем соответствие заголовков
            headers_in_file = excel_data[0] if excel_data else []
            header_matches = []

            for i in range(len(fixed_headers)):
                file_header = str(headers_in_file[i]) if i < len(headers_in_file) else ''
                expected = fixed_headers[i]

                # Сравниваем нормализованные заголовки
                file_normalized = file_header.strip().lower()
                expected_normalized = expected.strip().lower()
                matches = file_normalized == expected_normalized

                header_matches.append({
                    'expected': expected,
                    'actual': file_header,
                    'matches': matches
                })

            # Формируем предпросмотр
            preview_data = []

            # Добавляем заголовки с указанием соответствия
            mapped_headers = []
            for i, header_info in enumerate(header_matches):
                status = "✅" if header_info['matches'] else "❌"
                actual = header_info['actual'] if header_info['actual'] else f"Колонка {i + 1}"
                mapped_headers.append(f"{status} {actual} → {header_info['expected']}")

            preview_data.append(mapped_headers)

            # Добавляем данные (только первые 8 колонок)
            data_start = 1
            data_rows = excel_data[data_start:data_start + limit]

            for row_idx, row in enumerate(data_rows, start=2):
                preview_row = []
                for i in range(8):  # Теперь 8 фиксированных колонок
                    if i < len(row):
                        value = row[i]
                        # Обрезаем длинные значения для предпросмотра
                        if value and len(str(value)) > 100:
                            value = str(value)[:97] + "..."
                        preview_row.append(value)
                    else:
                        preview_row.append('')

                # Добавляем номер строки в начале
                preview_row.insert(0, row_idx - 1)  # -1 потому что первая строка - заголовки
                preview_data.append(preview_row)

            # Определяем реальное количество строк в файле
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
                ws = wb.active
                total_data_rows = ws.max_row - 1  # минус заголовок
                wb.close()
            except:
                # Если не получилось определить, используем количество прочитанных строк
                total_data_rows = len(excel_data) - 1 if len(excel_data) > 1 else 0

            correct_headers = sum(1 for h in header_matches if h['matches'])

            return True, {
                'preview': preview_data,
                'total_rows': total_data_rows,  # реальное количество строк
                'shown_rows': len(data_rows),  # показанное количество строк
                'expected_headers': fixed_headers,
                'actual_headers': headers_in_file,
                'header_matches': header_matches,
                'headers_correct': correct_headers,
                'headers_total': len(fixed_headers),
                'mapped_headers': mapped_headers,
                'file_name': Path(file_path).name
            }

        except Exception as e:
            print(f"Ошибка предпросмотра: {str(e)}")
            return False, str(e)


# Фабричная функция
def create_importer(json_file_path=None):
    """Создание экземпляра импортера"""
    return ExcelImporter(json_file_path)


if __name__ == "__main__":
    print("Тест импортера с фиксированными колонками")
    print("=" * 60)

    # Создаем тестовый импортер
    importer = create_importer()

    # Проверяем функцию очистки текста
    test_text = "Текст с _x000D_ Windows символами_x000D_и переносами\r\nстрок"
    cleaned = importer.clean_text(test_text)
    print(f"Тест очистки текста:")
    print(f"  Было: {repr(test_text)}")
    print(f"  Стало: {repr(cleaned)}")

    print("\nОжидаемые заголовки колонок:")
    expected = ['№', 'Вопрос', 'Ответ', 'Объяснение', 'Тема', 'Сложность', 'Скрытый', 'Версия']
    for i, header in enumerate(expected):
        print(f"  {i+1}. {header}")

    print("\nГотов к импорту файлов Excel!")
